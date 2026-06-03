from __future__ import annotations

import csv
import io
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = APP_DIR / "templates" / "wigwam_report_template.xlsx"
REPORT_SHEET = "Amazon Weekly Sales Rpt "
BUSINESS_SHEET = "Sheet2"
GOBROS_SHEET = "Sheet6"
DETAIL_FIRST_ROW = 3
DETAIL_LAST_ROW = 463

BUSINESS_COLUMNS = [
    "(Parent) ASIN",
    "(Child) ASIN",
    "Title",
    "SKU",
    "Sessions - Total",
    "Sessions - Total - B2B",
    "Session Percentage - Total",
    "Session Percentage - Total - B2B",
    "Page Views - Total",
    "Page Views - Total - B2B",
    "Page Views Percentage - Total",
    "Page Views Percentage - Total - B2B",
    "Featured Offer (Buy Box) Percentage",
    "Featured Offer (Buy Box) Percentage - B2B",
    "Units Ordered",
    "Units Ordered - B2B",
    "Unit Session Percentage",
    "Unit Session Percentage - B2B",
    "Ordered Product Sales",
    "Ordered Product Sales - B2B",
    "Total Order Items",
    "Total Order Items - B2B",
]


def generate_report(
    business_report: Path,
    inventory_report: Path,
    output_path: Path,
    month_ending: str | None = None,
    dealer_name: str = "GoBros",
    gobros_sales: Path | None = None,
) -> dict[str, Any]:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing template workbook: {TEMPLATE_PATH}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, output_path)

    wb = load_workbook(output_path)
    report_ws = wb[REPORT_SHEET]
    business_ws = wb[BUSINESS_SHEET]
    gobros_ws = wb[GOBROS_SHEET]

    month_value = _parse_month(month_ending)
    _refresh_business_sheet(business_ws, _read_table(business_report))
    _refresh_gobros_sheet(gobros_ws, _read_table(gobros_sales) if gobros_sales else [])

    inventory_rows = _read_table(inventory_report)
    inventory_by_sku = {
        str(row.get("sku", "")).strip(): row
        for row in inventory_rows
        if str(row.get("sku", "")).strip()
    }

    stats = _refresh_report_sheet(
        report_ws=report_ws,
        inventory_by_sku=inventory_by_sku,
        dealer_name=dealer_name,
        month_value=month_value,
    )

    try:
        wb.properties.title = "Wigwam Monthly Sales Report"
        wb.properties.subject = "Wigwam report generated from monthly Amazon exports"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)
    return {
        "output": str(output_path),
        "business_rows": max(business_ws.max_row - 1, 0),
        "inventory_rows": len(inventory_rows),
        "matched_inventory_rows": stats["matched_inventory_rows"],
        "missing_inventory_rows": stats["missing_inventory_rows"],
    }


def _refresh_business_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _clear_range(ws, 1, max(ws.max_row, len(rows) + 1), 1, 34)
    for col_idx, header in enumerate(BUSINESS_COLUMNS, start=1):
        ws.cell(1, col_idx).value = header

    helper_headers = {
        27: "SKU",
        28: "Units Ordered",
        29: "SKU",
        30: "Ordered Product Sales",
        31: "SKU",
        32: "Sessions - Total",
        33: "SKU",
        34: "Featured Offer (Buy Box) Percentage",
    }
    for col_idx, header in helper_headers.items():
        ws.cell(1, col_idx).value = header

    for row_idx, row in enumerate(rows, start=2):
        normalized = {_normalize_header(k): v for k, v in row.items()}
        values = [_business_value(normalized, col) for col in BUSINESS_COLUMNS]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx).value = value

        sku = values[3]
        units = values[14]
        sales = values[18]
        sessions = values[4]
        buy_box = values[12]
        helper_values = [sku, units, sku, sales, sku, sessions, sku, buy_box]
        for offset, value in enumerate(helper_values, start=27):
            ws.cell(row_idx, offset).value = value


def _refresh_gobros_sheet(ws, rows: list[dict[str, Any]]) -> None:
    _clear_range(ws, 1, max(ws.max_row, len(rows) + 1), 1, 17)
    headers = [
        "Product variant SKU",
        "Net items sold",
        "Product variant SKU",
        "Net sales",
        None,
        None,
        None,
        "Product title",
        "Product variant title",
        "Product variant SKU",
        "Net items sold",
        "Gross sales",
        "Discounts",
        "Returns",
        "Net sales",
        "Taxes",
        "Total sales",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = header

    for row_idx, row in enumerate(rows, start=2):
        normalized = {_normalize_header(k): v for k, v in row.items()}
        sku = _first_present(normalized, ["productvariantsku", "sku", "variant sku"])
        net_items = _number(_first_present(normalized, ["netitemssold", "net items sold", "quantity"]))
        net_sales = _number(_first_present(normalized, ["netsales", "net sales", "sales"]))
        source_values = [
            _first_present(normalized, ["producttitle", "product title", "title"]),
            _first_present(normalized, ["productvarianttitle", "product variant title", "variant title"]),
            sku,
            net_items,
            _number(_first_present(normalized, ["grosssales", "gross sales"])),
            _number(_first_present(normalized, ["discounts"])),
            _number(_first_present(normalized, ["returns"])),
            net_sales,
            _number(_first_present(normalized, ["taxes"])),
            _number(_first_present(normalized, ["totalsales", "total sales"])),
        ]
        for col_idx, value in enumerate([sku, net_items, sku, net_sales], start=1):
            ws.cell(row_idx, col_idx).value = value
        for col_idx, value in enumerate(source_values, start=8):
            ws.cell(row_idx, col_idx).value = value


def _refresh_report_sheet(report_ws, inventory_by_sku, dealer_name, month_value):
    matched = 0
    missing = 0
    for row_idx in range(DETAIL_FIRST_ROW, DETAIL_LAST_ROW + 1):
        fbm_sku = str(report_ws.cell(row_idx, 5).value or "").strip()
        fba_sku = str(report_ws.cell(row_idx, 6).value or "").strip()
        if not fbm_sku and not fba_sku:
            continue

        report_ws.cell(row_idx, 1).value = dealer_name
        if month_value is not None:
            report_ws.cell(row_idx, 2).value = month_value

        fbm_inventory = inventory_by_sku.get(fbm_sku, {})
        fba_inventory = inventory_by_sku.get(fba_sku, {})
        if fbm_inventory or fba_inventory:
            matched += 1
        else:
            missing += 1

        asin = fba_inventory.get("asin") or fbm_inventory.get("asin")
        if asin:
            report_ws.cell(row_idx, 7).value = asin

        report_ws.cell(row_idx, 18).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet6!$A$2:$B$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 19).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet6!$C$2:$D$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 21).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet2!$AA$1:$AB$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 22).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet2!$AC$1:$AD$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 23).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet2!$AE$1:$AF$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 24).value = f"=IFERROR(VLOOKUP(E{row_idx},Sheet2!$AG$1:$AH$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 26).value = f"=IFERROR(VLOOKUP(F{row_idx},Sheet2!$AA$1:$AB$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 27).value = f"=IFERROR(VLOOKUP(F{row_idx},Sheet2!$AC$1:$AD$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 28).value = f"=IFERROR(VLOOKUP(F{row_idx},Sheet2!$AE$1:$AF$20000,2,FALSE),0)"
        report_ws.cell(row_idx, 29).value = f"=IFERROR(VLOOKUP(F{row_idx},Sheet2!$AG$1:$AH$20000,2,FALSE),0)"

        report_ws.cell(row_idx, 31).value = _number(fbm_inventory.get("mfn-fulfillable-quantity"))
        report_ws.cell(row_idx, 32).value = _number(fba_inventory.get("afn-fulfillable-quantity"))
        inbound = (
            _number(fba_inventory.get("afn-inbound-working-quantity"))
            + _number(fba_inventory.get("afn-inbound-shipped-quantity"))
            + _number(fba_inventory.get("afn-inbound-receiving-quantity"))
        )
        report_ws.cell(row_idx, 33).value = inbound
        report_ws.cell(row_idx, 34).value = _number(fba_inventory.get("afn-unsellable-quantity"))
        report_ws.cell(row_idx, 35).value = _number(fba_inventory.get("afn-total-quantity"))
        report_ws.cell(row_idx, 36).value = f"=SUM(AE{row_idx}+AI{row_idx})"
        report_ws.cell(row_idx, 37).value = f"=SUM(AJ{row_idx}*O{row_idx})"

    report_ws.cell(465, 18).value = "=SUM(R3:R463)"
    report_ws.cell(465, 19).value = "=SUM(S3:S463)"
    report_ws.cell(465, 21).value = "=SUM(U3:U463)"
    report_ws.cell(465, 22).value = "=SUM(V3:V463)"
    report_ws.cell(465, 23).value = "=SUM(W3:W463)"
    report_ws.cell(465, 24).value = "=AVERAGE(X3:X463)"
    report_ws.cell(465, 26).value = "=SUM(Z3:Z463)"
    report_ws.cell(465, 27).value = "=SUM(AA3:AA463)"
    report_ws.cell(465, 28).value = "=SUM(AB3:AB463)"
    report_ws.cell(465, 29).value = "=AVERAGE(AC3:AC463)"
    for col in range(31, 40):
        letter = report_ws.cell(465, col).column_letter
        report_ws.cell(465, col).value = f"=SUM({letter}3:{letter}463)"

    return {"matched_inventory_rows": matched, "missing_inventory_rows": missing}


def _read_table(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel_tab if sample.count("\t") > sample.count(",") else csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(row) for row in reader]


def _business_value(row: dict[str, Any], header: str) -> Any:
    value = _first_present(row, [_normalize_header(header), header])
    if header in {"(Parent) ASIN", "(Child) ASIN", "Title", "SKU"}:
        return value
    if "Percentage" in header or "Session Percentage" in header:
        return _percent(value)
    return _number(value)


def _first_present(row: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        normalized = _normalize_header(key)
        if normalized in row and row[normalized] not in (None, ""):
            return row[normalized]
    return None


def _normalize_header(header: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(header or "").strip().lower())


def _number(value: Any) -> float | int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value).strip().replace("$", "").replace(",", "")
    if cleaned.endswith("%"):
        return _percent(cleaned)
    if cleaned in {"", "-", "--"}:
        return 0
    try:
        number = float(cleaned)
    except ValueError:
        return 0
    return int(number) if number.is_integer() else number


def _percent(value: Any) -> float:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value).strip().replace(",", "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
        try:
            return float(cleaned) / 100
        except ValueError:
            return 0
    try:
        return float(cleaned)
    except ValueError:
        return 0


def _parse_month(value: str | None):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return value


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None
